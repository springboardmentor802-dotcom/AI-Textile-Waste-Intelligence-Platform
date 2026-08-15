import io
import json
from datetime import datetime

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy import text

from .database import db_session, init_db, USING_POSTGRES
from . import auth
from . import notifications as notif
from .auth import (
    hash_password, verify_password, create_access_token, create_refresh_token,
    decode_token, get_current_user, require_roles, ROLES,
)
from .schemas import (
    RegisterRequest, LoginRequest, RefreshRequest, InventoryCreate, InventoryUpdate,
    UpdateProfileRequest, ChangePasswordRequest, NotificationPreferences,
)
from . import cv_analysis
from .material_classifier import classify_material
from . import scoring
from . import mongo

app = FastAPI(title="Textile Waste Intelligence Platform API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup():
    init_db()


@app.get("/api/health")
def health():
    return {"status": "ok", "database": "postgresql" if USING_POSTGRES else "sqlite (fallback)"}


# ================================================================= AUTH ====
@app.post("/api/auth/register", response_model=None)
def register(payload: RegisterRequest):
    if payload.role not in ROLES:
        raise HTTPException(400, f"Role must be one of {ROLES}")
    with db_session() as conn:
        existing = conn.execute(text("SELECT id FROM users WHERE email=:e"), {"e": payload.email}).first()
        if existing:
            raise HTTPException(400, "Email already registered")
        conn.execute(
            text("INSERT INTO users (full_name, email, hashed_password, role) VALUES (:n,:e,:p,:r)"),
            {"n": payload.full_name, "e": payload.email, "p": hash_password(payload.password), "r": payload.role},
        )
    return {"message": "Registered successfully. Please log in."}


@app.post("/api/auth/login")
def login(payload: LoginRequest):
    with db_session() as conn:
        row = conn.execute(
            text("SELECT id, full_name, hashed_password, role FROM users WHERE email=:e"),
            {"e": payload.email},
        ).mappings().first()
    if not row or not verify_password(payload.password, row["hashed_password"]):
        raise HTTPException(401, "Invalid email or password")

    access = create_access_token(row["id"], row["role"])
    refresh = create_refresh_token(row["id"])
    return {"access_token": access, "refresh_token": refresh, "token_type": "bearer",
            "role": row["role"], "full_name": row["full_name"]}


# OAuth2-compatible login (so a standard OAuth2PasswordRequestForm / Swagger "Authorize" button also works)
@app.post("/api/auth/token")
def login_oauth2(form_data: OAuth2PasswordRequestForm = Depends()):
    return login(LoginRequest(email=form_data.username, password=form_data.password))


@app.post("/api/auth/refresh")
def refresh_token(payload: RefreshRequest):
    data = decode_token(payload.refresh_token)
    if data.get("type") != "refresh":
        raise HTTPException(401, "Invalid refresh token")
    with db_session() as conn:
        row = conn.execute(
            text("SELECT id, role FROM users WHERE id=:id"), {"id": int(data["sub"])}
        ).mappings().first()
    if not row:
        raise HTTPException(401, "User not found")
    return {"access_token": create_access_token(row["id"], row["role"]), "token_type": "bearer"}


@app.post("/api/auth/logout")
def logout(payload: RefreshRequest, user=Depends(get_current_user)):
    with db_session() as conn:
        conn.execute(text("UPDATE refresh_tokens SET revoked=1 WHERE token=:t"), {"t": payload.refresh_token})
    return {"message": "Logged out"}


@app.get("/api/auth/profile")
def profile(user=Depends(get_current_user)):
    return user


@app.put("/api/auth/profile")
def update_profile(payload: UpdateProfileRequest, user=Depends(get_current_user)):
    fields = {k: v for k, v in payload.dict().items() if v is not None}
    if not fields:
        return {"message": "Nothing to update"}
    if "email" in fields:
        with db_session() as conn:
            clash = conn.execute(
                text("SELECT id FROM users WHERE email=:e AND id != :id"),
                {"e": fields["email"], "id": user["id"]},
            ).first()
        if clash:
            raise HTTPException(400, "Email already in use")
    set_clause = ", ".join([f"{k}=:{k}" for k in fields])
    with db_session() as conn:
        conn.execute(text(f"UPDATE users SET {set_clause} WHERE id=:id"), {**fields, "id": user["id"]})
    return {"message": "Profile updated"}


@app.post("/api/auth/change-password")
def change_password(payload: ChangePasswordRequest, user=Depends(get_current_user)):
    with db_session() as conn:
        row = conn.execute(
            text("SELECT hashed_password FROM users WHERE id=:id"), {"id": user["id"]}
        ).mappings().first()
        if not verify_password(payload.current_password, row["hashed_password"]):
            raise HTTPException(400, "Current password is incorrect")
        conn.execute(
            text("UPDATE users SET hashed_password=:p WHERE id=:id"),
            {"p": hash_password(payload.new_password), "id": user["id"]},
        )
    return {"message": "Password changed successfully"}


# ================================================================ ADMIN ====
@app.get("/api/admin/users")
def list_users(user=Depends(require_roles("admin"))):
    with db_session() as conn:
        rows = conn.execute(text("SELECT id, full_name, email, role, created_at FROM users ORDER BY id DESC")).mappings().all()
    return [dict(r) for r in rows]


@app.put("/api/admin/users/{user_id}/role")
def update_user_role(user_id: int, payload: dict, user=Depends(require_roles("admin"))):
    new_role = payload.get("role")
    if new_role not in ROLES:
        raise HTTPException(400, f"Role must be one of {ROLES}")
    with db_session() as conn:
        conn.execute(text("UPDATE users SET role=:r WHERE id=:id"), {"r": new_role, "id": user_id})
    return {"message": "Role updated"}


@app.delete("/api/admin/users/{user_id}")
def delete_user(user_id: int, user=Depends(require_roles("admin"))):
    if user_id == user["id"]:
        raise HTTPException(400, "Cannot delete your own account")
    with db_session() as conn:
        conn.execute(text("DELETE FROM users WHERE id=:id"), {"id": user_id})
    return {"message": "User deleted"}


@app.get("/api/dashboard/admin")
def admin_dashboard(user=Depends(require_roles("admin"))):
    with db_session() as conn:
        total_users = conn.execute(text("SELECT COUNT(*) c FROM users")).mappings().first()["c"]
        users_by_role = conn.execute(text("SELECT role, COUNT(*) c FROM users GROUP BY role")).mappings().all()
        total_analyses = conn.execute(text("SELECT COUNT(*) c FROM analyses")).mappings().first()["c"]
        total_inventory = conn.execute(text("SELECT COUNT(*) c FROM inventory")).mappings().first()["c"]
        total_notifications = conn.execute(text("SELECT COUNT(*) c FROM notifications")).mappings().first()["c"]
        recent_users = conn.execute(text("SELECT id, full_name, email, role, created_at FROM users ORDER BY id DESC LIMIT 5")).mappings().all()
        recent_analyses = conn.execute(text("SELECT id, filename, material, circularity_score, created_at FROM analyses ORDER BY id DESC LIMIT 5")).mappings().all()

    return {
        "total_users": total_users,
        "users_by_role": [dict(r) for r in users_by_role],
        "total_analyses": total_analyses,
        "total_inventory_records": total_inventory,
        "total_notifications": total_notifications,
        "recent_users": [dict(r) for r in recent_users],
        "recent_analyses": [dict(r) for r in recent_analyses],
        "system_status": "operational",
        "database_mode": "postgresql" if USING_POSTGRES else "sqlite (fallback)",
    }
@app.get("/api/dashboard/recycling-facility")
def recycling_facility_dashboard(user=Depends(require_roles("admin", "recycling_operator"))):
    with db_session() as conn:
        total_inventory = conn.execute(text("SELECT COUNT(*) c, SUM(quantity) q FROM inventory")).mappings().first()
        by_fabric = conn.execute(text("SELECT fabric_type, SUM(quantity) q, COUNT(*) c FROM inventory GROUP BY fabric_type")).mappings().all()
        recycling_opportunities = conn.execute(
            text("SELECT filename, material, circularity_score, recommendation FROM analyses WHERE circularity_score >= 65 ORDER BY circularity_score DESC LIMIT 10")
        ).mappings().all()
        by_recommendation = conn.execute(text("SELECT recommendation, COUNT(*) c FROM analyses GROUP BY recommendation")).mappings().all()
        avg_score = conn.execute(text("SELECT AVG(circularity_score) a FROM analyses")).mappings().first()["a"] or 0
        recovery_rate = conn.execute(
            text("SELECT COUNT(*) c FROM analyses WHERE waste_category IN ('Recyclable','Reusable','Upcyclable')")
        ).mappings().first()["c"]
        total_processed = conn.execute(text("SELECT COUNT(*) c FROM analyses")).mappings().first()["c"]

    return {
        "waste_inventory": {
            "total_batches": total_inventory["c"] or 0,
            "total_quantity_kg": total_inventory["q"] or 0,
            "by_fabric_type": [dict(r) for r in by_fabric],
        },
        "recycling_opportunities": [dict(r) for r in recycling_opportunities],
        "processing_analytics": {
            "by_recommendation": [dict(r) for r in by_recommendation],
            "average_circularity_score": round(avg_score, 1),
        },
        "recovery_statistics": {
            "recoverable_items": recovery_rate,
            "total_processed": total_processed,
            "recovery_rate_pct": round((recovery_rate / total_processed * 100), 1) if total_processed else 0,
        },
    }


@app.get("/api/dashboard/sustainability-manager")
def sustainability_manager_dashboard(user=Depends(require_roles("admin", "sustainability_manager"))):
    with db_session() as conn:
        rows = conn.execute(text("SELECT circularity_score, waste_category, created_at FROM analyses")).mappings().all()
        avg_score = conn.execute(text("SELECT AVG(circularity_score) a FROM analyses")).mappings().first()["a"] or 0

    if not rows:
        return {
            "sustainability_metrics": {"average_circularity_score": 0, "total_items_tracked": 0},
            "carbon_reduction_report": {"total_co2_saved_kg": 0},
            "waste_diversion_analytics": {"diversion_rate_pct": 0, "by_category": []},
            "esg_reporting": {"environmental_score": 0, "social_score": "N/A", "governance_score": "N/A"},
        }

    total_carbon = sum((r["circularity_score"] or 0) / 100 * 4.2 for r in rows)
    diverted = sum(1 for r in rows if r["waste_category"] != "Hazardous")
    category_counts = {}
    for r in rows:
        wc = r["waste_category"] or "Unspecified"
        category_counts[wc] = category_counts.get(wc, 0) + 1

    return {
        "sustainability_metrics": {
            "average_circularity_score": round(avg_score, 1),
            "total_items_tracked": len(rows),
        },
        "carbon_reduction_report": {
            "total_co2_saved_kg": round(total_carbon, 1),
        },
        "waste_diversion_analytics": {
            "diversion_rate_pct": round(diverted / len(rows) * 100, 1),
            "by_category": [{"waste_category": k, "count": v} for k, v in category_counts.items()],
        },
        "esg_reporting": {
            "environmental_score": round(avg_score, 1),
            "social_score": "N/A -- requires HR/labor data integration",
            "governance_score": "N/A -- requires compliance data integration",
        },
    }


@app.get("/api/dashboard/manufacturer")
def manufacturer_dashboard(user=Depends(require_roles("admin", "manufacturer"))):
    with db_session() as conn:
        production_waste = conn.execute(text("SELECT material, COUNT(*) c, AVG(circularity_score) a FROM analyses GROUP BY material")).mappings().all()
        by_recyclability = conn.execute(text("SELECT recyclability, COUNT(*) c FROM analyses GROUP BY recyclability")).mappings().all()
        avg_score = conn.execute(text("SELECT AVG(circularity_score) a FROM analyses")).mappings().first()["a"] or 0
        total = conn.execute(text("SELECT COUNT(*) c FROM analyses")).mappings().first()["c"]

    return {
        "production_waste_analysis": [dict(r) for r in production_waste],
        "circular_economy_insights": {
            "by_recyclability": [dict(r) for r in by_recyclability],
            "total_items_analyzed": total,
        },
        "material_recovery_reports": [dict(r) for r in production_waste],
        "sustainability_performance": {
            "average_circularity_score": round(avg_score, 1),
        },
    }

DEFAULT_SETTINGS = {
    "waste_collection_alerts": True,
    "recycling_opportunity_notifications": True,
    "sustainability_milestone_alerts": True,
    "inventory_warnings": True,
    "platform_announcements": True,
}


@app.get("/api/settings")
def get_settings(user=Depends(get_current_user)):
    with db_session() as conn:
        row = conn.execute(
            text("SELECT * FROM user_settings WHERE user_id=:id"), {"id": user["id"]}
        ).mappings().first()
    if not row:
        return DEFAULT_SETTINGS
    return {k: bool(v) for k, v in dict(row).items() if k != "user_id"}


@app.put("/api/settings")
def update_settings(payload: NotificationPreferences, user=Depends(get_current_user)):
    with db_session() as conn:
        existing = conn.execute(text("SELECT user_id FROM user_settings WHERE user_id=:id"), {"id": user["id"]}).first()
        data = payload.dict()
        if existing:
            set_clause = ", ".join([f"{k}=:{k}" for k in data])
            conn.execute(text(f"UPDATE user_settings SET {set_clause} WHERE user_id=:id"), {**data, "id": user["id"]})
        else:
            cols = ", ".join(["user_id"] + list(data.keys()))
            vals = ", ".join([":user_id"] + [f":{k}" for k in data])
            conn.execute(text(f"INSERT INTO user_settings ({cols}) VALUES ({vals})"), {**data, "user_id": user["id"]})
    return {"message": "Settings saved"}


# ========================================================= NOTIFICATIONS ====
@app.get("/api/notifications")
def list_notifications(unread_only: bool = False, user=Depends(get_current_user)):
    with db_session() as conn:
        query = "SELECT * FROM notifications WHERE user_id=:uid"
        if unread_only:
            query += " AND is_read=0" if not USING_POSTGRES else " AND is_read=FALSE"
        query += " ORDER BY created_at DESC LIMIT 50"
        rows = conn.execute(text(query), {"uid": user["id"]}).mappings().all()
    return [dict(r) for r in rows]


@app.get("/api/notifications/unread-count")
def unread_notification_count(user=Depends(get_current_user)):
    with db_session() as conn:
        query = "SELECT COUNT(*) c FROM notifications WHERE user_id=:uid AND is_read="
        query += "0" if not USING_POSTGRES else "FALSE"
        count = conn.execute(text(query), {"uid": user["id"]}).mappings().first()["c"]
    return {"unread_count": count}


@app.put("/api/notifications/{notification_id}/read")
def mark_notification_read(notification_id: int, user=Depends(get_current_user)):
    with db_session() as conn:
        conn.execute(
            text("UPDATE notifications SET is_read=:val WHERE id=:id AND user_id=:uid"),
            {"val": True, "id": notification_id, "uid": user["id"]},
        )
    return {"message": "Marked as read"}


@app.put("/api/notifications/mark-all-read")
def mark_all_notifications_read(user=Depends(get_current_user)):
    with db_session() as conn:
        conn.execute(
            text("UPDATE notifications SET is_read=:val WHERE user_id=:uid"),
            {"val": True, "uid": user["id"]},
        )
    return {"message": "All notifications marked as read"}


@app.delete("/api/notifications/{notification_id}")
def delete_notification(notification_id: int, user=Depends(get_current_user)):
    with db_session() as conn:
        conn.execute(
            text("DELETE FROM notifications WHERE id=:id AND user_id=:uid"),
            {"id": notification_id, "uid": user["id"]},
        )
    return {"message": "Notification deleted"}


# ============================================================ INVENTORY ====
@app.get("/api/inventory")
def list_inventory(search: str = "", user=Depends(get_current_user)):
    with db_session() as conn:
        if search:
            rows = conn.execute(
                text("SELECT * FROM inventory WHERE batch_id ILIKE :s OR fabric_type ILIKE :s ORDER BY id DESC"
                     if USING_POSTGRES else
                     "SELECT * FROM inventory WHERE batch_id LIKE :s OR fabric_type LIKE :s ORDER BY id DESC"),
                {"s": f"%{search}%"},
            ).mappings().all()
        else:
            rows = conn.execute(text("SELECT * FROM inventory ORDER BY id DESC")).mappings().all()
    return [dict(r) for r in rows]


@app.post("/api/inventory")
def create_inventory(payload: InventoryCreate,
                      user=Depends(require_roles("admin", "manufacturer", "recycling_operator"))):
    with db_session() as conn:
        existing = conn.execute(text("SELECT id FROM inventory WHERE batch_id=:b"), {"b": payload.batch_id}).first()
        if existing:
            raise HTTPException(400, "Batch ID already exists")
        conn.execute(
            text("""INSERT INTO inventory (batch_id, fabric_type, source, quantity, color, condition,
                    collection_date, created_by) VALUES (:batch_id,:fabric_type,:source,:quantity,:color,
                    :condition,:collection_date,:created_by)"""),
            {**payload.dict(), "created_by": user["id"]},
        )
        notif.notify_waste_collection_alert(conn, user["id"], payload.batch_id, payload.fabric_type, payload.quantity)
    return {"message": "Inventory record created"}


@app.put("/api/inventory/{item_id}")
def update_inventory(item_id: int, payload: InventoryUpdate,
                      user=Depends(require_roles("admin", "manufacturer", "recycling_operator"))):
    fields = {k: v for k, v in payload.dict().items() if v is not None}
    if not fields:
        return {"message": "Nothing to update"}
    set_clause = ", ".join([f"{k}=:{k}" for k in fields])
    with db_session() as conn:
        conn.execute(text(f"UPDATE inventory SET {set_clause} WHERE id=:id"), {**fields, "id": item_id})
    return {"message": "Updated"}


@app.delete("/api/inventory/{item_id}")
def delete_inventory(item_id: int, user=Depends(require_roles("admin", "recycling_operator"))):
    with db_session() as conn:
        conn.execute(text("DELETE FROM inventory WHERE id=:id"), {"id": item_id})
    return {"message": "Deleted"}


# ================================================================ ANALYZE ====
@app.post("/api/analyze-batch")
async def analyze_batch(files: list[UploadFile] = File(...), user=Depends(get_current_user)):
    if len(files) > 5:
        raise HTTPException(400, "Maximum 5 images per batch")

    results = []
    for f in files:
        start = datetime.utcnow()
        image_bytes = await f.read()
        try:
            img = cv_analysis.load_image(image_bytes)
        except ValueError as e:
            results.append({"filename": f.filename, "error": str(e)})
            continue

        h, w = img.shape[:2]

        color_result = cv_analysis.analyze_color(img)
        texture_result = cv_analysis.analyze_texture(img)
        damage_result = cv_analysis.analyze_damage(img)
        contamination_result = cv_analysis.analyze_contamination(img)
        material_result = classify_material(img, color_result, texture_result)

        scores = scoring.compute_scores(material_result, texture_result, damage_result, contamination_result)
        waste_result = scoring.classify_waste(material_result, damage_result, contamination_result, scores)
        recycling_result = scoring.recommend_recycling(material_result, waste_result, scores, damage_result, contamination_result)
        sustainability_result = scoring.sustainability_assessment(material_result, scores)
        environmental_result = scoring.environmental_impact(sustainability_result, scores)

        processing_time_ms = round((datetime.utcnow() - start).total_seconds() * 1000, 1)

        result = {
            "filename": f.filename,
            "image_info": {"width": w, "height": h, "size_bytes": len(image_bytes),
                            "processing_time_ms": processing_time_ms},
            "material_classification": material_result,
            "texture_analysis": texture_result,
            "color_analysis": color_result,
            "damage_detection": damage_result,
            "contamination_detection": contamination_result,
            "waste_classification": waste_result,
            "recycling_recommendation": recycling_result,
            "sustainability_assessment": sustainability_result,
            "environmental_impact": environmental_result,
            "scores": scores,
        }

        mongo_doc_id = mongo.save_analysis_document(result)
        result["mongo_doc_id"] = mongo_doc_id

        with db_session() as conn:
            conn.execute(
                text("""INSERT INTO analyses (filename, material, confidence, fiber_composition, recyclability,
                        circularity_score, waste_category, recommendation, mongo_doc_id, created_by)
                        VALUES (:filename,:material,:confidence,:fiber,:recyclability,:circ,:waste,:rec,:mongo_id,:uid)"""),
                {
                    "filename": f.filename,
                    "material": material_result["material"],
                    "confidence": material_result["confidence"],
                    "fiber": material_result["fiber_composition"],
                    "recyclability": material_result["recyclability"],
                    "circ": scores["circularity_score"],
                    "waste": waste_result["waste_category"],
                    "rec": recycling_result["best_recommendation"],
                    "mongo_id": mongo_doc_id,
                    "uid": user["id"],
                },
            )
            notif.notify_recycling_opportunity(conn, user["id"], f.filename, scores["circularity_score"], recycling_result["best_recommendation"])
            notif.notify_inventory_warning(conn, user["id"], f.filename, waste_result["waste_category"])
            total = conn.execute(text("SELECT COUNT(*) c FROM analyses WHERE created_by=:uid"), {"uid": user["id"]}).mappings().first()["c"]
            avg_circ = conn.execute(text("SELECT AVG(circularity_score) a FROM analyses WHERE created_by=:uid"), {"uid": user["id"]}).mappings().first()["a"] or 0
            notif.notify_sustainability_milestone(conn, user["id"], total, avg_circ)

        results.append(result)

    return {"count": len(results), "results": results}


@app.post("/api/analyze")
async def analyze_image(file: UploadFile = File(...), user=Depends(get_current_user)):
    start = datetime.utcnow()
    image_bytes = await file.read()
    try:
        img = cv_analysis.load_image(image_bytes)
    except ValueError as e:
        raise HTTPException(400, str(e))

    h, w = img.shape[:2]

    color_result = cv_analysis.analyze_color(img)
    texture_result = cv_analysis.analyze_texture(img)
    damage_result = cv_analysis.analyze_damage(img)
    contamination_result = cv_analysis.analyze_contamination(img)
    material_result = classify_material(img, color_result, texture_result)

    scores = scoring.compute_scores(material_result, texture_result, damage_result, contamination_result)
    waste_result = scoring.classify_waste(material_result, damage_result, contamination_result, scores)
    recycling_result = scoring.recommend_recycling(material_result, waste_result, scores, damage_result, contamination_result)
    sustainability_result = scoring.sustainability_assessment(material_result, scores)
    environmental_result = scoring.environmental_impact(sustainability_result, scores)

    processing_time_ms = round((datetime.utcnow() - start).total_seconds() * 1000, 1)

    result = {
        "filename": file.filename,
        "image_info": {"width": w, "height": h, "size_bytes": len(image_bytes),
                        "processing_time_ms": processing_time_ms},
        "material_classification": material_result,
        "texture_analysis": texture_result,
        "color_analysis": color_result,
        "damage_detection": damage_result,
        "contamination_detection": contamination_result,
        "waste_classification": waste_result,
        "recycling_recommendation": recycling_result,
        "sustainability_assessment": sustainability_result,
        "environmental_impact": environmental_result,
        "scores": scores,
    }

    mongo_doc_id = mongo.save_analysis_document(result)
    result["mongo_doc_id"] = mongo_doc_id

    with db_session() as conn:
        conn.execute(
            text("""INSERT INTO analyses (filename, material, confidence, fiber_composition, recyclability,
                    circularity_score, waste_category, recommendation, mongo_doc_id, created_by)
                    VALUES (:filename,:material,:confidence,:fiber,:recyclability,:circ,:waste,:rec,:mongo_id,:uid)"""),
            {
                "filename": file.filename,
                "material": material_result["material"],
                "confidence": material_result["confidence"],
                "fiber": material_result["fiber_composition"],
                "recyclability": material_result["recyclability"],
                "circ": scores["circularity_score"],
                "waste": waste_result["waste_category"],
                "rec": recycling_result["best_recommendation"],
                "mongo_id": mongo_doc_id,
                "uid": user["id"],
            },
        )
        notif.notify_recycling_opportunity(conn, user["id"], file.filename, scores["circularity_score"], recycling_result["best_recommendation"])
        notif.notify_inventory_warning(conn, user["id"], file.filename, waste_result["waste_category"])
        total = conn.execute(text("SELECT COUNT(*) c FROM analyses WHERE created_by=:uid"), {"uid": user["id"]}).mappings().first()["c"]
        avg_circ = conn.execute(text("SELECT AVG(circularity_score) a FROM analyses WHERE created_by=:uid"), {"uid": user["id"]}).mappings().first()["a"] or 0
        notif.notify_sustainability_milestone(conn, user["id"], total, avg_circ)

    return result


@app.get("/api/history")
def history(user=Depends(get_current_user)):
    with db_session() as conn:
        rows = conn.execute(text("SELECT * FROM analyses ORDER BY id DESC LIMIT 200")).mappings().all()
    return [dict(r) for r in rows]


# ================================================================ REPORTS ====
@app.get("/api/reports")
def list_reports(material: str = "", waste_category: str = "", user=Depends(get_current_user)):
    query = "SELECT * FROM analyses WHERE 1=1"
    params = {}
    if material:
        query += " AND material = :material"
        params["material"] = material
    if waste_category:
        query += " AND waste_category = :waste_category"
        params["waste_category"] = waste_category
    query += " ORDER BY id DESC LIMIT 500"
    with db_session() as conn:
        rows = conn.execute(text(query), params).mappings().all()
    return [dict(r) for r in rows]


@app.get("/api/reports/summary")
def reports_summary(user=Depends(get_current_user)):
    with db_session() as conn:
        total = conn.execute(text("SELECT COUNT(*) c FROM analyses")).mappings().first()["c"]
        avg_circ = conn.execute(text("SELECT AVG(circularity_score) a FROM analyses")).mappings().first()["a"] or 0
        by_waste = conn.execute(text("SELECT waste_category, COUNT(*) c FROM analyses GROUP BY waste_category")).mappings().all()
        by_material = conn.execute(text("SELECT material, COUNT(*) c FROM analyses GROUP BY material")).mappings().all()
    return {
        "total_reports": total,
        "average_circularity_score": round(avg_circ, 1),
        "by_waste_category": [dict(r) for r in by_waste],
        "by_material": [dict(r) for r in by_material],
    }

@app.get("/api/reports/export/excel")
def export_reports_excel(user=Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    with db_session() as conn:
        rows = conn.execute(text("SELECT * FROM analyses ORDER BY id DESC")).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Waste Analysis Report"

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for r in rows:
            ws.append([str(v) if v is not None else "" for v in dict(r).values()])
        for col in ws.columns:
            max_len = max(len(str(c.value)) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(30, max_len + 2)
    else:
        ws.append(["No analysis records yet."])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=textile_waste_report.xlsx"},
    )

@app.get("/api/reports/export/csv")
def export_reports_csv(user=Depends(get_current_user)):
    import csv

    with db_session() as conn:
        rows = conn.execute(text("SELECT * FROM analyses ORDER BY id DESC")).mappings().all()

    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        for r in rows:
            writer.writerow(dict(r))
    else:
        buf.write("No analysis records yet.\n")

    byte_buf = io.BytesIO(buf.getvalue().encode("utf-8"))
    return StreamingResponse(byte_buf, media_type="text/csv",
                              headers={"Content-Disposition": "attachment; filename=textile_waste_report.csv"})


# =============================================================== DASHBOARD ====
@app.get("/api/dashboard/summary")
def dashboard_summary(user=Depends(get_current_user)):
    with db_session() as conn:
        total_uploads = conn.execute(text("SELECT COUNT(*) c FROM analyses")).mappings().first()["c"]
        total_inventory = conn.execute(text("SELECT COUNT(*) c FROM inventory")).mappings().first()["c"]
        avg_circularity = conn.execute(text("SELECT AVG(circularity_score) a FROM analyses")).mappings().first()["a"] or 0
        materials = conn.execute(text("SELECT material, COUNT(*) c FROM analyses GROUP BY material")).mappings().all()
        fabric_dist = conn.execute(text("SELECT fabric_type, SUM(quantity) q FROM inventory GROUP BY fabric_type")).mappings().all()

    return {
        "total_uploads": total_uploads,
        "materials_classified": total_uploads,
        "inventory_records": total_inventory,
        "recycling_score": round(avg_circularity, 1),
        "carbon_saving_estimate_kg": round(avg_circularity * total_uploads * 0.04, 1) if total_uploads else 0,
        "waste_diversion_pct": round(min(98, avg_circularity + 8), 1) if avg_circularity else 0,
        "circularity_score": round(avg_circularity, 1),
        "recovery_rate_pct": round(avg_circularity, 1),
        "material_distribution": [dict(m) for m in materials],
        "fabric_type_distribution": [dict(f) for f in fabric_dist],
    }


@app.get("/api/report/batch/pdf")
def download_batch_report(ids: str, user=Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    id_list = [int(i) for i in ids.split(",") if i.strip()]
    if not id_list:
        raise HTTPException(400, "No analysis ids provided")

    with db_session() as conn:
        rows = []
        for analysis_id in id_list:
            row = conn.execute(text("SELECT * FROM analyses WHERE id=:id"), {"id": analysis_id}).mappings().first()
            if row:
                rows.append(row)

    if not rows:
        raise HTTPException(404, "No matching analyses found")

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    for idx, row in enumerate(rows):
        y = height - 60
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, y, "Textile Waste Intelligence Platform - Analysis Report")
        y -= 20
        c.setFont("Helvetica", 9)
        c.drawString(50, y, f"Item {idx + 1} of {len(rows)}")
        y -= 30
        c.setFont("Helvetica", 11)
        for label, value in [
            ("File", row["filename"]),
            ("Material", row["material"]),
            ("Confidence", f"{row['confidence']}%"),
            ("Fiber Composition", row["fiber_composition"]),
            ("Recyclability", row["recyclability"]),
            ("Circularity Score", f"{row['circularity_score']}/100"),
            ("Waste Category", row["waste_category"]),
            ("Recommendation", row["recommendation"]),
            ("Generated", str(row["created_at"])),
        ]:
            c.drawString(50, y, f"{label}: {value}")
            y -= 22
        c.showPage()

    c.save()
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": "attachment; filename=analysis_batch_report.pdf"})


@app.get("/api/report/{analysis_id}/pdf")
def download_report(analysis_id: int, user=Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    with db_session() as conn:
        row = conn.execute(text("SELECT * FROM analyses WHERE id=:id"), {"id": analysis_id}).mappings().first()
    if not row:
        raise HTTPException(404, "Analysis not found")

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 60

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, "Textile Waste Intelligence Platform - Analysis Report")
    y -= 30
    c.setFont("Helvetica", 11)
    for label, value in [
        ("File", row["filename"]),
        ("Material", row["material"]),
        ("Confidence", f"{row['confidence']}%"),
        ("Fiber Composition", row["fiber_composition"]),
        ("Recyclability", row["recyclability"]),
        ("Circularity Score", f"{row['circularity_score']}/100"),
        ("Waste Category", row["waste_category"]),
        ("Recommendation", row["recommendation"]),
        ("Generated", str(row["created_at"])),
    ]:
        c.drawString(50, y, f"{label}: {value}")
        y -= 22

    c.showPage()
    c.save()
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": f"attachment; filename=analysis_{analysis_id}.pdf"})


@app.get("/api/sustainability/report/pdf")
def sustainability_report_pdf(user=Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    with db_session() as conn:
        rows = conn.execute(text(
            "SELECT circularity_score, recyclability, waste_category, recommendation FROM analyses"
        )).mappings().all()
        avg = conn.execute(text("SELECT AVG(circularity_score) a FROM analyses")).mappings().first()["a"] or 0

    total = len(rows)
    recyclability_score_map = {"High": 90, "Medium": 60, "Low": 30}
    total_carbon = total_water = total_recovery = 0.0
    for r in rows:
        base = (r["circularity_score"] or 0) / 100
        total_carbon += base * 4.2
        total_water += base * 2650
        total_recovery += recyclability_score_map.get(r["recyclability"], 50)
    avg_recovery = (total_recovery / total) if total else 0
    benchmark = scoring.benchmark_against_industry(avg)

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 60

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, "Textile Waste Intelligence Platform")
    y -= 22
    c.setFont("Helvetica-Bold", 13)
    c.drawString(50, y, "Sustainability Report")
    y -= 30
    c.setFont("Helvetica", 11)
    for label, value in [
        ("Total Items Analyzed", total),
        ("Total Carbon Saved (kg CO2)", round(total_carbon, 1)),
        ("Total Water Saved (Liters)", round(total_water, 0)),
        ("Average Resource Recovery (%)", round(avg_recovery, 1)),
        ("Average Circularity Score", round(avg, 1)),
        ("Industry Benchmark Score", benchmark["industry_benchmark_score"]),
        ("Best-in-Class Reference Score", benchmark["top_performer_score"]),
        ("Standing", benchmark["standing"]),
    ]:
        c.drawString(50, y, f"{label}: {value}")
        y -= 22

    c.showPage()
    c.save()
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": "attachment; filename=sustainability_report.pdf"})

# ========================================================= SUSTAINABILITY ====
@app.get("/api/sustainability/summary")
def sustainability_summary(user=Depends(get_current_user)):
    with db_session() as conn:
        rows = conn.execute(text(
            "SELECT circularity_score, recyclability, waste_category, recommendation, created_at FROM analyses"
        )).mappings().all()

    if not rows:
        return {
            "total_carbon_saved_kg": 0, "total_water_saved_liters": 0,
            "avg_waste_diversion_pct": 0, "avg_circularity_score": 0,
            "avg_resource_recovery_pct": 0, "trend": [],
            "recommendation_distribution": [], "waste_category_distribution": [],
        }

    recyclability_score_map = {"High": 90, "Medium": 60, "Low": 30}
    total_carbon = total_water = total_diversion = total_recovery = 0.0
    trend = []
    rec_counts = {}
    waste_counts = {}

    for r in rows:
        score = r["circularity_score"] or 0
        base = score / 100
        carbon = round(base * 4.2, 2)
        water = round(base * 2650, 0)
        diversion = min(98, score + 8)
        recovery = recyclability_score_map.get(r["recyclability"], 50)

        total_carbon += carbon
        total_water += water
        total_diversion += diversion
        total_recovery += recovery

        trend.append({"date": str(r["created_at"])[:10], "circularity_score": score})

        rec = r["recommendation"] or "Unspecified"
        rec_counts[rec] = rec_counts.get(rec, 0) + 1

        wc = r["waste_category"] or "Unspecified"
        waste_counts[wc] = waste_counts.get(wc, 0) + 1

    n = len(rows)
    return {
        "total_carbon_saved_kg": round(total_carbon, 1),
        "total_water_saved_liters": round(total_water, 0),
        "avg_waste_diversion_pct": round(total_diversion / n, 1),
        "avg_circularity_score": round(sum(r["circularity_score"] or 0 for r in rows) / n, 1),
        "avg_resource_recovery_pct": round(total_recovery / n, 1),
        "trend": trend,
        "recommendation_distribution": [{"recommendation": k, "count": v} for k, v in rec_counts.items()],
        "waste_category_distribution": [{"waste_category": k, "count": v} for k, v in waste_counts.items()],
    }