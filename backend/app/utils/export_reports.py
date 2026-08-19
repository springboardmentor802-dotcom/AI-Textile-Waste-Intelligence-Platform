import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


# ==========================================
# 📊 1. EXCEL REPORT GENERATOR (OpenPyXL)
# ==========================================

def generate_excel_report(records: list) -> io.BytesIO:
    """Generates an executive-styled Excel spreadsheet from inventory/scan records."""
    output = io.BytesIO()
    
    if records:
        df = pd.DataFrame(records)
    else:
        df = pd.DataFrame([{"Message": "No inventory or scan records available"}])

    df.columns = [str(col).replace("_", " ").title() for col in df.columns]

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Sustainability Audit")
        
        workbook = writer.book
        worksheet = writer.sheets["Sustainability Audit"]
        
        # Styles
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Slate 900
        header_font = Font(name="Arial", size=11, bold=True, color="10B981") # Emerald 500
        data_font = Font(name="Arial", size=10, color="334155")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        # Header Row Formatting
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Data Rows Formatting
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-fit Column Widths
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 15)

    output.seek(0)
    return output


# ==========================================
# 📄 2. SIMPLE INVENTORY PDF SUMMARY
# ==========================================

def generate_pdf_text_summary(records: list) -> io.BytesIO:
    """Generates a formatted PDF inventory summary using ReportLab."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=letter,
        rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=12
    )
    normal_style = styles['Normal']

    story = []
    story.append(Paragraph("Textile Sustainability Inventory Summary", title_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#10B981'), spaceAfter=15))

    if not records:
        story.append(Paragraph("No inventory records found.", normal_style))
    else:
        table_data = [["Batch ID", "Fabric Type", "Quantity (Kg)", "Condition", "Status"]]
        for rec in records:
            table_data.append([
                str(rec.get('batch_id', 'N/A')),
                str(rec.get('fabric_type', 'N/A')),
                f"{rec.get('quantity', 0)} Kg",
                str(rec.get('condition', 'N/A')),
                str(rec.get('status', 'Active'))
            ])

        t = Table(table_data, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 1.3*inch, 1.2*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        story.append(t)

    doc.build(story)
    output.seek(0)
    return output


# ==========================================
# 🚀 3. MULTI-ENGINE AI PDF REPORT BUILDER
# ==========================================

def generate_multi_engine_pdf_report(batch_id: str, results: dict) -> io.BytesIO:
    """Generates an Executive-Grade PDF Report covering all 7 AI Processing Engines."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()

    # Custom ReportLab Typography
    brand_title = ParagraphStyle(
        'BrandTitle',
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=4
    )
    sub_title = ParagraphStyle(
        'SubTitle',
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.HexColor('#10B981'),
        spaceAfter=12
    )
    engine_heading = ParagraphStyle(
        'EngineHeading',
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor('#0F172A'),
        spaceBefore=10,
        spaceAfter=6
    )
    cell_key = ParagraphStyle(
        'CellKey',
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.HexColor('#475569')
    )
    cell_val = ParagraphStyle(
        'CellVal',
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#0F172A')
    )

    story = []

    # 1. Header Section
    story.append(Paragraph("AI TEXTILE WASTE INTELLIGENCE PLATFORM", brand_title))
    story.append(Paragraph(f"COMPREHENSIVE MULTI-ENGINE DIAGNOSTIC REPORT | REF: {batch_id}", sub_title))
    story.append(HRFlowable(width="100%", thickness=2.5, color=colors.HexColor('#10B981'), spaceAfter=14))

    # 2. Engines Mapping
    engines = [
        ("1. TEXTILE IMAGE ANALYSIS ENGINE", results.get("image_analysis_engine", {})),
        ("2. MATERIAL CLASSIFICATION ENGINE", results.get("material_classification_engine", {})),
        ("3. TEXTILE WASTE CLASSIFICATION ENGINE", results.get("waste_classification_engine", {})),
        ("4. RECYCLING RECOMMENDATION ENGINE", results.get("recycling_recommendation_engine", {})),
        ("5. SUSTAINABILITY INTELLIGENCE ENGINE", results.get("sustainability_intelligence_engine", {})),
        ("6. ENVIRONMENTAL IMPACT ASSESSMENT ENGINE", results.get("environmental_impact_engine", {})),
        ("7. WASTE SCORING ENGINE (5-TIER WEIGHTED MODEL)", results.get("waste_scoring_engine", {}))
    ]

    for title, engine_data in engines:
        story.append(Paragraph(title, engine_heading))

        if isinstance(engine_data, dict) and engine_data:
            table_rows = []
            keys = list(engine_data.keys())

            # Render key-value pairs in a 2-column grid
            for i in range(0, len(keys), 2):
                k1 = keys[i]
                v1 = str(engine_data[k1]) if engine_data[k1] is not None else "N/A"
                k1_fmt = k1.replace("_", " ").title()

                if i + 1 < len(keys):
                    k2 = keys[i+1]
                    v2 = str(engine_data[k2]) if engine_data[k2] is not None else "N/A"
                    k2_fmt = k2.replace("_", " ").title()

                    table_rows.append([
                        Paragraph(k1_fmt, cell_key),
                        Paragraph(v1, cell_val),
                        Paragraph(k2_fmt, cell_key),
                        Paragraph(v2, cell_val)
                    ])
                else:
                    table_rows.append([
                        Paragraph(k1_fmt, cell_key),
                        Paragraph(v1, cell_val),
                        Paragraph("", cell_key),
                        Paragraph("", cell_val)
                    ])

            engine_table = Table(
                table_rows,
                colWidths=[1.8*inch, 2.0*inch, 1.8*inch, 2.0*inch]
            )
            engine_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(engine_table)
        else:
            story.append(Paragraph("No engine log data returned for this section.", cell_val))

        story.append(Spacer(1, 8))

    doc.build(story)
    output.seek(0)
    return output