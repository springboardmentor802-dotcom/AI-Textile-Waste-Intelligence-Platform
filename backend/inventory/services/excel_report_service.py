import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)


def generate_waste_report_excel(report_data):
    """
    Builds a single-item waste report as an .xlsx workbook,
    mirroring the same data used in the PDF report.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Waste Report"

    rows = [
        ("Section", "Field", "Value"),
        ("Image Analysis", "Brightness", report_data["image_analysis"]["brightness_analysis"].get("brightness_level", "")),
        ("Image Analysis", "Texture", report_data["image_analysis"]["texture_analysis"].get("texture_level", "")),
        ("Image Analysis", "Contamination Suspected", report_data["image_analysis"]["damage_contamination_check"].get("contamination_suspected", "")),
        ("Material Classification", "Predicted Fiber Type", report_data["material_classification"].get("predicted_fiber_type", "")),
        ("Material Classification", "Confidence (%)", report_data["material_classification"].get("confidence", "")),
        ("Waste Categorization", "Waste Category", report_data["waste_categorization"].get("waste_category", "")),
        ("Waste Categorization", "Reason", report_data["waste_categorization"].get("reason", "")),
        ("Recyclability Assessment", "Circularity Score", report_data["recyclability_assessment"].get("circularity_score", "")),
        ("Recyclability Assessment", "Circularity Category", report_data["recyclability_assessment"].get("circularity_category", "")),
    ]

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = BODY_FONT

    widths = [26, 26, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_inventory_excel(textile_queryset):
    """
    Exports the full textile inventory (list of TextileWaste batches) as .xlsx.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        "Batch ID", "Material Type", "Quantity (kg)", "Color", "Source",
        "Condition", "Status", "Detected Material", "Circularity Score",
        "Waste Category", "Collection Date", "Date Added",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(textile_queryset, start=2):
        values = [
            item.batch_id, item.material_type, item.quantity, item.color,
            item.source, item.condition, item.status, item.detected_material,
            item.circularity_score, item.waste_category,
            str(item.collection_date) if item.collection_date else "",
            str(item.date_added),
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value).font = BODY_FONT

    widths = [16, 14, 12, 12, 16, 14, 14, 16, 16, 20, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer